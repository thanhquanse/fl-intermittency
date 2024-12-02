#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from openpyxl import load_workbook, Workbook

def extract_first_110_columns(input_file, output_file):
    # Load the workbook and the first sheet
    wb = load_workbook(input_file)
    source_sheet = wb.active  # Assuming the first sheet is the source

    # Create a new workbook for the output
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Extracted Columns"

    # Copy the first 110 columns
    max_columns = min(110, source_sheet.max_column)  # Ensure no more than available columns are copied
    for col_idx in range(1, max_columns + 1):  # Columns are 1-indexed
        for row_idx, cell in enumerate(source_sheet.iter_rows(min_col=col_idx, max_col=col_idx), start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=cell[0].value)

    # Save the new workbook
    new_wb.save(output_file)
    print(f"First {max_columns} columns extracted and saved to {output_file}")


# In[ ]:

def move_and_append_columns_to_24_hour_sheets(input_file, output_file):
    # Load the workbook and the first sheet
    wb = load_workbook(input_file)
    source_sheet = wb.active  # Assuming the first sheet is the source

    # Get the first column data (reference column)
    reference_column = [cell.value for cell in source_sheet['A']]

    # Prepare column names for 24 hours (Hour 0 to Hour 23)
    hour_columns = [f"Hour {i}" for i in range(24)]

    # Process the first 100 columns starting from the second column
    for col_index, col_cells in enumerate(source_sheet.iter_cols(min_col=2, max_col=101), start=2):
        # Flatten column data and skip None values
        column_data = [cell.value for cell in col_cells if cell.value is not None]

        # Create a new sheet for the current column
        sheet_name = f"Sheet_{col_index}"
        new_sheet = wb.create_sheet(title=sheet_name)

        # Write the reference column to column A of the new sheet
        for row_idx, value in enumerate(reference_column, start=1):
            new_sheet.cell(row=row_idx, column=1, value=value)

        # Write hour column names (Hour 0 to Hour 23) to the first row
        for hour_idx, hour_name in enumerate(hour_columns, start=2):
            new_sheet.cell(row=1, column=hour_idx, value=hour_name)

        # Append data from the current column into respective hour columns in 24-row chunks
        current_row = 2
        for chunk_start in range(0, len(column_data), 24):
            chunk = column_data[chunk_start:chunk_start + 24]
            for hour_idx, value in enumerate(chunk):
                new_sheet.cell(row=current_row, column=hour_idx + 2, value=value)
            current_row += 1

    # Save the workbook with the new sheets
    wb.save(output_file)
    print(f"Data moved and appended into sheets and saved to {output_file}")

# In[ ]:

def extract_first_x_rows(input_file, output_file):
    # Load the workbook
    wb = load_workbook(input_file)

    # Create a new workbook for the output
    new_wb = Workbook()
    if "Sheet" in new_wb.sheetnames:
        new_wb.remove(new_wb["Sheet"])  # Remove the default sheet

    # Process each sheet starting from the second
    for sheet_index, sheet in enumerate(wb.worksheets[1:], start=2):  # Skip the first sheet
        # Create a new sheet in the output workbook
        new_sheet = new_wb.create_sheet(title=sheet.title)

        # Extract the first 1,097 rows from the current sheet
        max_rows = min(2000000, sheet.max_row)  # Ensure no more than available rows are copied
        for row_idx in range(1, max_rows + 1):
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)

    # Save the new workbook
    new_wb.save(output_file)
    print(f"First 10000 rows from each sheet (starting from the second) extracted and saved to {output_file}")


# In[ ]:

def move_columns_to_new_sheets(input_file, output_file):
    # Load the workbook and the first sheet
    wb = load_workbook(input_file)
    source_sheet = wb.active  # Assuming the first sheet is the source

    # Get the first column data (reference column)
    reference_column = [cell.value for cell in source_sheet['A']]

    # Prepare column names for 24 hours (Hour 0 to Hour 23)
    hour_columns = [f"Hour {i}" for i in range(24)]

    # Process the first 100 columns starting from the second column
    for col_index, col_cells in enumerate(source_sheet.iter_cols(min_col=2, max_col=101), start=2):
        # Flatten column data and skip None values
        column_data = [cell.value for cell in col_cells if cell.value is not None]

        # Split the column data into chunks of 24 rows
        for chunk_index in range(0, len(column_data), 24):
            # Create a new sheet for each chunk
            sheet_name = f"Sheet_{col_index}_Chunk_{chunk_index // 24 + 1}"
            new_sheet = wb.create_sheet(title=sheet_name)

            # Write the reference column to column A of the new sheet
            for row_idx, value in enumerate(reference_column, start=1):
                new_sheet.cell(row=row_idx, column=1, value=value)

            # Write hour column names (Hour 0 to Hour 23) to the first row
            for hour_idx, hour_name in enumerate(hour_columns, start=2):
                new_sheet.cell(row=1, column=hour_idx, value=hour_name)

            # Write data from the current chunk into respective hour columns
            for row_offset, value in enumerate(column_data[chunk_index:chunk_index + 24]):
                new_sheet.cell(row=2, column=row_offset + 2, value=value)

    # Save the workbook with the new sheets
    wb.save(output_file)
    print(f"Data distributed into sheets and saved to {output_file}")

def distribute_sheet_values(infile, outfile):
    # Load the workbook
    file_path = infile  # Replace with your file path
    workbook = load_workbook(file_path)

    # Loop through each sheet in the workbook
    for sheet in workbook.worksheets:
        # Get all values from the second column (column B) starting from row 2
        second_column_values = [row[1].value for row in sheet.iter_rows(min_row=2, max_col=2) if row[1].value is not None]
        
        # Create 24 new columns (Hour_0, Hour_1, ..., Hour_23)
        for i in range(24):
            sheet.cell(row=1, column=i + 3, value=f"Hour_{i}")  # Write headers starting from column 3 (C)

        # Append values to the 24 columns in chunks of 24 rows
        start_row = 2
        for i in range(0, len(second_column_values), 24):
            chunk = second_column_values[i:i+24]  # Get the next 24 rows
            for col_index, value in enumerate(chunk, start=3):  # Start writing in column C
                sheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1

    # Save the updated workbook
    workbook.save(outfile)


def update_cell_value(file_path):
    # Load the Excel workbook
    workbook = load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Collect all values in column B starting from B3
        values = []
        for row in range(3, sheet.max_row + 1):  # Start from B3
            cell_value = sheet[f'B{row}'].value
            if isinstance(cell_value, (int, float)):  # Ensure it's numeric
                values.append(cell_value)
        
        # Calculate the mean value
        mean_value = sum(values) / len(values) if values else 0  # Avoid division by zero

        # Replace the value of cell B2 with the mean
        sheet['B2'] = mean_value

    # Save the workbook
    workbook.save(file_path)
    print(f"Updated cell B2 in all sheets of {file_path}.")


if __name__ == "__main__":
    # for dataset in ['weather', 'traffic' ,'train', 'psm']:
    for dataset in ['m4']:
        # extract_first_110_columns(f"{dataset}/{dataset}.xlsx", f"{dataset}/{dataset}_1.xlsx")
        # move_and_append_columns_to_24_hour_sheets(f"{dataset}/{dataset}_1.xlsx", f"{dataset}/{dataset}_2.xlsx")
        distribute_sheet_values(f"{dataset}/{dataset}.xlsx", f"{dataset}/{dataset}_2.xlsx")
        # extract_first_x_rows(f"{dataset}/{dataset}_2.xlsx", f"{dataset}/standard_{dataset}.xlsx")
        # move_columns_to_new_sheets(f"{dataset}/{dataset}_3.xlsx", f"{dataset}/standard_{dataset}.xlsx")
        # update_cell_value(f"{dataset}/standard_{dataset}.xlsx")
